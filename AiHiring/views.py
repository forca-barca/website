from django.http import HttpResponse
from django.template import loader
from django.shortcuts import render
from django.core.files.storage import FileSystemStorage
from AiHiring.scratch_10 import *
from django.http import FileResponse
from openpyxl import Workbook
from django.conf import settings
#import nltk
#from nltk.corpus import stopwords 
import re
import difflib
import hashlib
from AiHiring.models import RegisterLogin,UserOTP
from django.core.exceptions import ObjectDoesNotExist
import smtplib
import datetime as dt
from datetime import timedelta
from random import randint
from email.mime.text import MIMEText
import pytz
import os
from wsgiref.util import FileWrapper
import mimetypes
def index(request):
	return render(request,'AiHiring/index.html')

def examples(request):
	return render(request,'AiHiring/examples.html')

myfile = []
uploaded_file_url = ''
data = ''
fileNameList = []
fileDataList = []
#import glob

#files = glob.glob(settings.BASE_DIR+"\media/")
#for f in files:
#	os.remove(f)
def page(request):
	if request.method == 'POST' and request.FILES['uploadFOLDER']:
		myfile = request.FILES.getlist('uploadFOLDER')
		path=settings.BASE_DIR+"\media/"
		import shutil
		shutil.rmtree(path)
		for files in myfile:
			data = ""
			#print(type(myfile))
			fs = FileSystemStorage()
			filename = fs.save(files.name, files)
			pdflist = ['pdf','PDF','Pdf','pDf','pdF','PDf','PdF','pDF']
			print(filename.split('.')[1])
			if filename.split('.')[1] in pdflist:
				uploaded_file_url = fs.url(filename)
				#print(uploaded_file_url.split('/')[2])
			
				finalPath = settings.BASE_DIR+uploaded_file_url
				print(finalPath)
				#print(settings.BASE_DIR+"\media/")

				if (' ' in filename) == True:
					finalPath = settings.BASE_DIR +"\media/"+ re.sub(r"\s+", "", filename, flags=re.UNICODE)
					os.rename(os.path.join(path,filename),os.path.join(path,re.sub(r"\s+", "", filename, flags=re.UNICODE)))
					fileNameList.append(re.sub(r"\s+", "", filename, flags=re.UNICODE))
				else:
					fileNameList.append(uploaded_file_url.split('/')[2])
				data = convert(finalPath)
				fileDataList.append(data)
				print(len(fileDataList),len(fileNameList))
				#word_tokens=input_file_words(data,[])
				#request.session['word_tokens']=word_tokens
				#request.session['data']=data
		request.session['fileData']=fileDataList
		request.session['fileName']=fileNameList

			#print(word_tokens[0])
			#print(len(word_tokens))
			#f = open(finalPath,'r')
			#data += f.read()
			#print(data)
	return render(request,'AiHiring/page.html')

def another_page(request):
	if request.method == 'POST' and request.POST.get('want_result')=='giveResult':
		jobLOCATION=request.session['jobLOCATION']
		cutoffJL=request.session['cutoffJL']
		jobSKILLS=request.session['jobSKILLS']
		cutoffJS=request.session['cutoffJS']
		university=request.session['university']
		cutoffUNIV=request.session['cutoffUNIV']
		companies=request.session['companies']
		cutoffCOMP=request.session['cutoffCOMP']
		degree=request.session['degree']
		cutoffDEG=request.session['cutoffDEG']
		major=request.session['major']
		cutoffMAJ=request.session['cutoffMAJ']

		avgCUTOFF = request.POST.get('averageCUTOFF')
		print('average cutoff::::::',avgCUTOFF)

		word_tokens=request.session['word_tokens']
		data=request.session['data']
		sent_tokens=input_file_lines(data,[])
		nltkWORDtokens = word_tokenize(data)
		new_data = str(preprocess(data))
		new_tokens=str(sent_tokenize(new_data))
		new_words=str(word_tokenize(new_data))
		new_tokens_lines=input_file_lines(new_data,[])

		if int(avgCUTOFF)>=50:
			msg="This resume is selected..."
			currentLocation=location_matched(nltkWORDtokens,jobLOCATION)
			skills=programming_skills(nltkWORDtokens,jobSKILLS)
			university=collegeScore(sent_tokens,university)
			company=companies_matched(data,companies)
			qualification=degree_matched(word_tokens,sent_tokens,degree,major,data)

		else:
			msg="This resume is not selected..."
			currentLocation=None
			skills=None
			university=None
			company=None
			qualification=None
		return render(request,'AiHiring/another_page.html',{'msg' : msg , 'cL' : currentLocation , 'sk' : skills , 'univ' : university , 'comp' : company , 'qual' : qualification} )



	if request.POST.get('hehe'):
			print(request.POST.get('hehe'))	

	

	


	if request.method == 'POST' and request.POST.get('detail_submitted')=='Next Step':
	#	if avgCUTOFF in locals():
	#		avgCUTOFF = request.POST.get('averageCUTOFF')
		#	print('averageCUTOFF :::',averageCUTOFF)
		cutOFFlist=[]
		weightList=[]
		gege = request.POST.get('detail_submitted')
		print("detail_submitted :::::::::::::::::::::::::::::::::::::::::::::::",gege)


		jobLOCATION = request.POST.getlist('jobLOCATION')
		#if jobLOCATION != request.session['jobLOCATION']:
		#	jobLOCATION=request.session['jobLOCATION']

		cutoffJL = request.POST.get('cutoffJL')
		cutOFFlist.append(cutoffJL)
		#if cutoffJL != request.session['cutoffJL']:
		#	cutoffJL=request.session['cutoffJL']

		weightJL = request.POST.get('weightJL')
		if weightJL!='':
			weightList.append(int(weightJL))
		else:
			weightList.append(int(0))
		
		jobSKILLS = request.POST.getlist('jobSKILLS')
		#if jobSKILLS != request.session['jobSKILLS']:
		#	jobSKILLS=request.session['jobSKILLS']
		
		cutoffJS = request.POST.get('cutoffJS')
		cutOFFlist.append(cutoffJS)
		#if cutoffJS != request.session['cutoffJS']:
		#	cutoffJS=request.session['cutoffJS']

		weightJS = request.POST.get('weightJS')
		if weightJS!='':
			weightList.append(int(weightJS))
		else:
			weightList.append(int(0))

		companies = request.POST.getlist('companies')
		#if companies != request.session['companies']:
		#	companies=request.session['companies']
		
		cutoffCOMP = request.POST.get('cutoffCOMP')
		cutOFFlist.append(cutoffCOMP)
		#if cutoffCOMP != request.session['cutoffCOMP']:
		#	cutoffCOMP=request.session['cutoffCOMP']

		weightCOMP = request.POST.get('weightCOMP')
		if weightCOMP!='':
			weightList.append(int(weightCOMP))
		else:
			weightList.append(int(0))


		university = request.POST.getlist('university')
		#if university != request.session['university']:
		#	university=request.session['university']
		
		cutoffUNIV = request.POST.get('cutoffUNIV')
		cutOFFlist.append(cutoffUNIV)
		#if cutoffUNIV != request.session['cutoffUNIV']:
		#	cutoffUNIV=request.session['cutoffUNIV']

		weightUNIV = request.POST.get('weightUNIV')
		if weightUNIV!='':
			weightList.append(int(weightUNIV))
		else:
			weightList.append(int(0))

		degree = request.POST.getlist('degree')
		#if degree != request.session['degree']:
		#	degree=request.session['degree']
		
		cutoffDEG = request.POST.get('cutoffDEG')
		cutOFFlist.append(cutoffDEG)
		#if cutoffDEG != request.session['cutoffDEG']:
		#	cutoffDEG=request.session['cutoffDEG']

		weightDEG = request.POST.get('weightDEG')
		if weightDEG!='':
			weightList.append(int(weightDEG))
		else:
			weightList.append(int(0))

		major = request.POST.getlist('major')
		#if major != request.session['major']:
		#	major=request.session['major']
		
		cutoffMAJ = request.POST.get('cutoffMAJ')
		cutOFFlist.append(cutoffMAJ)
		#if cutoffMAJ != request.session['cutoffMAJ']:
		#	cutoffMAJ=request.session['cutoffMAJ']

		weightMAJ = request.POST.get('weightMAJ')
		if weightMAJ!='':
			weightList.append(int(weightMAJ))
		else:
			weightList.append(int(0))

		print(weightList,"::::::::::::::::weightList")
		request.session['jobLOCATION']=jobLOCATION
		request.session['cutoffJL']=cutoffJL
		request.session['jobSKILLS']=jobSKILLS
		request.session['cutoffJS']=cutoffJS
		request.session['university']=university
		request.session['cutoffUNIV']=cutoffUNIV
		request.session['companies']=companies
		request.session['cutoffCOMP']=cutoffCOMP
		request.session['degree']=degree
		request.session['cutoffDEG']=cutoffDEG
		request.session['major']=major
		request.session['cutoffMAJ']=cutoffMAJ

		#print(jobLOCATION)
		#print(cutoffJL)
		#print(jobSKILLS)
		#print(cutoffJS)
		#print(university)
		#print(cutoffUNIV)
		#print(companies)
		#print(cutoffCOMP)
		#print(degree)
		#print(cutoffDEG)
		#print(major)
		#print(cutoffMAJ)
		#print("listttttttttttt:::::::::",cutOFFlist)


		#word_tokens=request.session['word_tokens']
		#data=request.session['data']
		#sent_tokens=input_file_lines(data,[])
		#nltkWORDtokens = word_tokenize(data)
		#new_data = str(preprocess(data))
		#new_tokens=str(sent_tokenize(new_data))
		#new_words=str(word_tokenize(new_data))
		#new_tokens_lines=input_file_lines(new_data,[])


		fileDataList=request.session['fileData']
		fileNameList=request.session['fileName']

		ulist=[]
		currLOCmatch=[]
		currLOCscore=[]
		skillMatch=[]
		skillScore=[]
		univMatch=[]
		univScore=[]
		comMatch=[]
		comScore=[]
		quaMatch=[]
		quaScore=[]
		emailList=[]
		phoneList=[]
		#print(len(weightJL),len(weightJS),len(weightUNIV),len(weightCOMP),len(weightDEG),len(weightMAJ)	,"suiiiiiiiiiiiiii")
		if(len(weightJL)==0):
			weightJL=0
		if(len(weightJS)==0):
			weightJS=0
		if(len(weightUNIV)==0):
			weightUNIV=0
		if(len(weightCOMP)==0):
			weightCOMP=0
		if(len(weightDEG)==0):
			weightDEG=0
		if(len(weightMAJ)==0):
			weightMAJ=0



		averageScoreList=[]

		for _ in range(0,len(fileDataList)):
			sent_tokens=input_file_lines(fileDataList[_],[])
			nltkWORDtokens=word_tokenize(fileDataList[_])
			ulist=[]
			[ulist.append(x.lower()) for x in nltkWORDtokens if x.lower() not in ulist]
			word_tokens=input_file_words(fileDataList[_],[])

			new_data = str(preprocess(fileDataList[_]))
			new_tokens=str(sent_tokenize(new_data))
			new_words=str(word_tokenize(new_data))
			new_tokens_lines=input_file_lines(new_data,[])

			common_new=' '.join(unique_list(fileDataList[_].split()))
			common_new1=input_file_words(common_new,[])

			currentLocation=location_matched(ulist,jobLOCATION)
			currLOCmatch.append(currentLocation['locMatch'])
			currLOCscore.append(currentLocation['locScore'])
			
			skills=programming_skills(ulist,jobSKILLS)
			skillMatch.append(skills['skiMatch'])
			skillScore.append(skills['skiScore'])

			university=collegeScore(sent_tokens,university)
			univMatch.append(university['colMatch'])
			univScore.append(university['colScore'])

			company=companies_matched(fileDataList[_],companies)
			comMatch.append(company['compMatch'])
			comScore.append(company['compScore'])
			
			qualification=degree_matched(word_tokens,sent_tokens,degree,major,fileDataList[_])
			quaMatch.append(qualification['degMatch'])
			quaScore.append(qualification['degScore'])
			#print(type(qualification['degScore']))

			averageScore=overall_weightage_cal(float(currentLocation['locScore']),float(weightJL),float(skills['skiScore']),float(weightJS),float(university['colScore']),float(weightUNIV),float(company['compScore']),float(weightCOMP),float(qualification['degScore']),float(weightDEG))
			#print(averageScore)
			#skillsList.append(skills)
			#universityList.append(university)
			#companyList.append(company)
			#qualificationList.append(qualification)
			averageScoreList.append(averageScore)

			emailList.append(email_id(fileDataList[_]))
			phoneList.append(contact_no_mod(fileDataList[_]))
		#print(skillMatch)
		print(emailList,phoneList)
		summ=0
		newLEN=0
		#for _i in range(len(cutOFFlist)):
		#	if len(cutOFFlist[_i])>0:
		#		summ+=int(cutOFFlist[_i])
		#		newLEN+=1
		defaultScore=60
		if len(averageScoreList)!=0:
			avgScore=60
		else:
			avgScore=60
		#print(avgScore)
		iteration=[]
		for i in range(0,len(averageScoreList)):
			iteration.append(i)
		#print(type(iteration))

		




		
		#print(len(word_tokens))
		#print(len(data))
		#print(len(sent_tokens))
		#avgCutoff = average_score(int(cutoffJL),int(cutoffJS),cutoffUNIV,cutoffCOMP,cutoffDEG,cutoffMAJ)
		#currentLocation=location_matched(nltkWORDtokens,jobLOCATION)
		#skills=programming_skills(nltkWORDtokens,jobSKILLS)
		#university=collegeScore(sent_tokens,university)
		#company=companies_matched(data,companies)
		#qualification=degree_matched(word_tokens,sent_tokens,degree,major,data)
		#averageScore=float((currentLocation+skills+university+company+qualification)/5)*100
		#defaultScore=float((int(cutoffJL)+int(cutoffJS)+int(cutoffUNIV)+int(cutoffCOMP)+int(cutoffDEG)+int(cutoffMAJ))/6)
		parameters=[]
		if (request.POST.get('wantLOC')=='wantLOC'):
			parameters.append(1)
		else:
			parameters.append(0)
		if (request.POST.get('wantSKI')=='wantSKI'):
			parameters.append(1)
		else:
			parameters.append(0)
		if (request.POST.get('wantCOM')=='wantCOM'):
			parameters.append(1)
		else:
			parameters.append(0)
		if (request.POST.get('wantUNI')=='wantUNI'):
			parameters.append(1)
		else:
			parameters.append(0)
		if (request.POST.get('wantDEG')=='wantDEG'):
			parameters.append(1)
		else:
			parameters.append(0)
		if (request.POST.get('wantMAJ')=='wantMAJ'):
			parameters.append(1)
		else:
			parameters.append(0)
		print("parar::::::::::::",parameters)
		return render(request,'AiHiring/another_page.html',{'emailList':emailList,'phoneList':phoneList,'parameters':parameters,'weightList':weightList,'default' : defaultScore , 'avgScore': avgScore , 'fileNameList' : fileNameList , 'averageScoreList' : averageScoreList , 'loopCount' : iteration , 'cl' : currLOCscore , 'cl1' : currLOCmatch, 'sk' : skillScore, 'sk1' : skillMatch , 'un' : univScore , 'un1' : univMatch , 'co' : comScore, 'co1' : comMatch , 'qu' : quaScore,'qu1' : quaMatch} )

	

def contact(request):
	return render(request,'AiHiring/contact.html')

def get(request):
	return render(request, 'AiHiring/contact.html')

def post(request):
	return render(request, 'AiHiring/contact.html')


def fL(request):
	if request.method == 'POST' and request.POST.get('save')=='save':
		print("save:::::::::::::::::",request.POST.get('save'))
		tempp=request.POST.get('myList')
		tempp1=request.POST.get('myList1')
		tempp2=request.POST.get('myList2')
		clean_tempp=''
		clean_tempp1=''
		clean_tempp2=''
		for _ in tempp:
			if _ == "[" or _ == "]" or _ == "'":
				continue
			else:
				clean_tempp+=_
		for _ in tempp1:
			if _ == "[" or _ == "]" or _ == "'":
				continue
			else:
				clean_tempp1+=_
		for _ in tempp2:
			if _ == "[" or _ == "]" or _ == "'":
				continue
			else:
				clean_tempp2+=_
		fileARR=clean_tempp.split(',')
		emailARR=clean_tempp1.split(',')
		phoneARR=clean_tempp2.split(',')
		print(clean_tempp,"::::::::::::")
		print(clean_tempp1,"::::::::::::::::::::::")
		print(clean_tempp2,":::::::::::::::::::::::::::::::::::::")
		print(tempp2)
		wb = Workbook()
		dest_filename = settings.BASE_DIR+'/finalRESUMEs.xlsx'
		ws = wb.active
		ws.title = "range names"
		indx=''
		indx1=''
		indx2=''
		for _ in range(len(fileARR)):
			indx='A'+str(_+1)
			indx1='E'+str(_+1)
			indx2='H'+str(_+1)
			ws[indx]=fileARR[_]
			ws[indx1]=emailARR[_]
			ws[indx2]=phoneARR[_]
		wb.save(filename = dest_filename)
		return render(request,'finalLIST.html',{'fileARR':fileARR})

def excel(request):
 file_path = settings.BASE_DIR+'/finalRESUMEs.xlsx'
 file_wrapper = FileWrapper(open(file_path,'rb'))
 file_mimetype = mimetypes.guess_type(file_path)
 response = HttpResponse(file_wrapper, content_type=file_mimetype )
 response['X-Sendfile'] = file_path
 response['Content-Length'] = os.stat(file_path).st_size
 response['Content-Disposition'] = 'attachment; filename=finalRESUMEs.xlsx'
 return response

def selected(request):
	locs = request.POST.get('locs')
	print(locs)
	locm=''
	old_locm = request.POST.get('locm')
	for _ in old_locm:
		if _!='[' and _!=']' and _!="'":
			locm+=_
	print(locm)
	skis = request.POST.get('skis')
	print(skis)
	skim=''
	old_skim = request.POST.get('skim')
	for _ in old_skim:
		if _!='[' and _!=']' and _!="'":
			skim+=_
	print(skim)
	unis = request.POST.get('unis')
	print(unis)
	unim=''
	old_unim = request.POST.get('unim')
	for _ in old_unim:
		if _!='[' and _!=']' and _!="'":
			unim+=_
	print(unim)
	coms = request.POST.get('coms')
	print(coms)
	comm=''
	old_comm = request.POST.get('comm')
	for _ in old_comm:
		if _!='[' and _!=']' and _!="'":
			comm+=_
	print(comm)
	quas = request.POST.get('quas')
	print(quas)
	quam=''
	old_quam = request.POST.get('quam')
	for _ in old_quam:
		if _!='[' and _!=']' and _!="'":
			quam+=_
	print(quam)
	avgs = request.POST.get('avgs')
	print(avgs)
	file = request.POST.get('file')
	print(file)
	emam=''
	old_emam = request.POST.get('emam')
	for _ in old_emam:
		if _!='[' and _!=']' and _!="'":
			emam+=_
	print(file)
	phom = request.POST.get('phom')
	print(phom)
	paraStr = request.POST.get('para')
	print(type(paraStr))
	para=[]
	paraAS=paraStr.split(',')
	print(type(paraAS[0]))
	para=[];
	for _ in range(len(paraAS)):
		para.append(int(paraAS[_]))
	return render(request, 'selected.html',{'para':para,'emam':emam,'phom':phom,'locs':locs,'locm':locm,'skis':skis,'skim':skim,'unis':unis,'unim':unim,'coms':coms,'comm':comm,'quas':quas,'quam':quam,'avgs':avgs,'file':file})

def hi(request):
	filename = request.POST.get('filename')
	print(filename)
	finalPath = settings.BASE_DIR + '/media/' + str(filename)
	return FileResponse(open(finalPath, 'rb'), content_type='application/pdf')

'''
def simple_upload(request):
    if request.method == 'POST' and request.FILES['uploadFOLDER']:
        myfile = request.FILES['uploadFOLDER']
        print(myfile)
        fs = FileSystemStorage()
        filename = fs.save(myfile.name, myfile)
        uploaded_file_url = fs.url(filename)
        return render(request, 'AiHiring/index.html', {'uploaded_file_url': uploaded_file_url})
    return render(request, 'AiHiring/index.html')



'''

def firstPage(request):
	return render(request, 'firstPage.html')

	

def signupin(request):
	if request.method == 'POST' and request.POST.get('pswdCHNG')=='pswdCHNG':
		#username=request.POST['username2']
		newpswd=request.POST['newpass']
		renewpswd=request.POST.get('renewpass')
		change=RegisterLogin.objects.get(username=request.session['username'])
		print(newpswd,"NEWWWWWW")
		#####hash_object = hashlib.md5(newpswd.encode('utf-8')).hexdigest()
		#print(hash_object.hexdigest())
		######RegisterLogin.objects.create(companyname=tem[0],username=tem[1],email=tem[2],password=hash_object.hexdigest())
		change.password = hashlib.md5(newpswd.encode('utf-8')).hexdigest()
		change.username = request.session['username']
		change.email = request.session["email"]
		change.companyname = "harsh"
		print("halellujahhhhhhhhhhhhhhhhhhhh")
		change.save()
		return render(request, 'signupin.html')
	else:
		return render(request, 'signupin.html')

def auth(request):
	if request.method == 'POST' and request.POST.get('login')=='login':
		tem = request.POST.get('username')
		try:
			#print("jjjjjjjjjjjjjjjjjjjjjjjjjjjjj")
			cheese_blog = RegisterLogin.objects.get(username=tem)
			print(cheese_blog,";;;;;;;;;;;;;;;;;;;;;")
			print(cheese_blog.password)
			if hashlib.md5(request.POST.get('password').encode('utf-8')).hexdigest()==cheese_blog.password and cheese_blog.activated==True:
				return render(request, 'AiHiring/examples.html')
			else:
				if hashlib.md5(request.POST.get('password').encode('utf-8')).hexdigest()!=cheese_blog.password:
					return render(request, 'signupin.html',{'impmsg':'Wrong Password. Please enter the correct password or use Forget password to change the password. Thank You.'})
				elif cheese_blog.activated==False:
					return render(request, 'signupin.html',{'impmsg':'Please activate your account. Link had been sent to your email. Thank You.'})	
		except:
			#print("jjjjjjjjjjjjjjjjjjjjjjjjjjjjj")
			return render(request, 'signupin.html',{'impmsg':'Please Register yourself. This username is not yet registered. Thank You.'})
	else:
		tem=[]
		tem.append(request.POST.get('compname1'))
		tem.append(request.POST.get('username1'))
		tem.append(request.POST.get('email1'))
		tem.append(request.POST.get('password1'))
		tem.append(request.POST.get('repassword1'))
		#print(request.POST['username1'])
		try:
			cheese_blog = RegisterLogin.objects.get(username=tem[1])
			return render(request, 'signupin.html',{'impmsg':'This Username already exist. Please Choose another Username and Register again. Thank You'})
		except:
			hash_object = hashlib.md5(tem[3].encode('utf-8'))
			#print(hash_object.hexdigest())
			RegisterLogin.objects.create(companyname=tem[0],username=tem[1],email=tem[2],password=hash_object.hexdigest())
			content = '<a href="http://127.0.0.1:8000/AiHiring/activate.html"></a>'
			sender = 'patelharsh.ab@gmail.com'
			receivers = tem[2]
			msg = MIMEText(u'Greetings from DataBytes Analytics Pvt. Ltd. Thank for using our Product. Please <a href="http://127.0.0.1:8000/AiHiring/activate.html?'+tem[1]+'">Click Here</a> to activate your account.','html')
			msg['Subject'] = 'AI Based Hiring - Account Activation Link.'
			msg['From'] = sender
			msg['To'] = receivers
			#s.sendmail(me, [you], msg.as_string())
			#print("Successfully sent email")
			#s.quit()
			print(content)
		
			print(tem[2],type(tem[2]))
			mail=smtplib.SMTP('smtp.gmail.com:587')
			mail.starttls()
			mail.login(sender,'Harsh15111995')
			#mail.sendmail(sender, receivers, content)   
			mail.sendmail(sender, receivers, msg.as_string())
			print("Successfully sent email")      
			#print("Successfully sent email")
			mail.close()
			return render(request, 'signupin.html',{'tem':tem})

def activate(request):
	print("i am in activate view////////////////////")
	print(str(request.get_full_path).split()[-1].split("'")[1].split('?')[-1])
	try:
		cheese_blog = RegisterLogin.objects.get(username=str(request.get_full_path).split()[-1].split("'")[1].split('?')[-1])
		cheese_blog.activated = True
		cheese_blog.save()
	except:
		return render(request, 'firstPage.html')
	print(request.GET)
	return render(request, 'signupin.html')

def forget(request):
	return render(request,'forget.html')

def otp(request):
	username=request.POST['username2']
	email=request.POST.get('email2')
	request.session['username']=username
	request.session["email"]=email
	print("request.session:::::::::::::",request.session['username'])
	print(username,'SSSSSSSSSSSSSSSSSSSSS',type(username))
	
	sender = 'patelharsh.ab@gmail.com'
	receivers = email
	
	
	
	
	#cheese_blog = RegisterLogin.objects.get(username=username)
	#print(cheese_blog)
	now = dt.datetime.now()
	delta = dt.timedelta(seconds = 1800)
	t = now.time()
	try:
		print("jjjjjjjjjjjjjjjjjjjjjjjjjjjjj")
		cheese_blog = RegisterLogin.objects.get(username=username)
		try:
			cheese_blog1 = UserOTP.objects.get(username=username)
			cheese_blog1.delete()
			def random_with_N_digits(n):
				range_start = 10**(n-1)
				range_end = (10**n)-1
				return randint(range_start, range_end)
			content = str(random_with_N_digits(4))
			mail=smtplib.SMTP('smtp.gmail.com:587')
			mail.starttls()
			mail.login(sender,'Harsh15111995')
			mail.sendmail(sender, receivers, content)         
			print("Successfully sent email")
			mail.close()
			UserOTP.objects.create(username=username,email=cheese_blog,otp=content,startTime=now,endTime=now + delta)
		except:
			def random_with_N_digits(n):
				range_start = 10**(n-1)
				range_end = (10**n)-1
				return randint(range_start, range_end)
			content = str(random_with_N_digits(4))
			mail=smtplib.SMTP('smtp.gmail.com:587')
			mail.starttls()
			mail.login(sender,'Harsh15111995')
			mail.sendmail(sender, receivers, content)         
			print("Successfully sent email")
			mail.close()
			UserOTP.objects.create(username=username,email=cheese_blog,otp=content,startTime=now,endTime=now + delta)
		print(cheese_blog,";;;;;;;;;;;;;;;;;;;;;")
		return render(request, 'otp.html')
	except ObjectDoesNotExist:
		print("hshshshshh")
		return render(request, 'forget.html',{'impmsg':'Please Register yourself. This username is not yet registered. Thank You.'})
	#print((dt.datetime.combine(dt.date(1,1,1),t) + delta).time())
	#print(datetime.datetime.time(datetime.datetime.now()),datetime.datetime.time(datetime.datetime.now())+timedelta(seconds=5))
		#UserOTP.objects.create(username=username,email=cheese_blog,otp=content,startTime=dt.datetime.time(dt.datetime.now()),endTime=(dt.datetime.combine(dt.date(1,1,1),t) + delta).time())
		#print(cheese_blog,";;;;;;;;;;;;;;;;;;;;;")
		#return render(request, 'AiHiring/index.html')
	#except:
	#	return render(request, 'signupin.html')
	#return render(request,'otp.html',{'usr':username,'em':email})


def changePSWD(request):
	usrOTP = request.POST.get('OTP')
	now = dt.datetime.now().replace(tzinfo=None)
	delta = dt.timedelta(seconds = 10)
	t = now.time()
	username=request.session['username']
	email=request.session["email"]
	cheese_blog = UserOTP.objects.get(username=username)
	strtTM = cheese_blog.startTime.replace(tzinfo=None)
	endTM = cheese_blog.endTime.replace(tzinfo=None)
	dbOTP = cheese_blog.otp
	#print(type(usrOTP),type(dbOTP))
	#print(t,strtTM,endTM)
	#if(now.time()>strtTM.time()):
	#	print(now.time(),strtTM.time())
	#if usrOTP==dbOTP:
	#	print("bcbcbcbcb")
	#if (now>strtTM and now<endTM):
	#	print('111111fvgbhjnkm')
	if (usrOTP==dbOTP and now>strtTM and now<endTM):
		print('222222fvgbhjnkm')
		return render(request, 'changePSWD.html',{'helo':request.session['username']})
	else:
		print('333333fvgbhjnkm')
		if usrOTP!=dbOTP:
			return render(request,'otp.html',{'impmsg':'You have entered the wrong OTP. Please Enter the correct OTP. Thank You.'})
		elif now>endTM:
			return render(request,'otp.html',{'impmsg':'This OTP has expired. Please use the new OTP. Thank You.'})
		else:
			return render(request,'otp.html',{'impmsg':'heheheehhe'})